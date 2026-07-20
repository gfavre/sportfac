import re

from django.db.models import Max
from django.db.models import Min
from django.utils.translation import gettext as _
from openpyxl import load_workbook

from profiles.models import SchoolYear

from .models import Teacher


TEACHER_CORRESPONDANCE_DICT = {
    "number": ["Numéro SPEV", "ID LAGAPEO"],
    "last_name": ["Nom"],
    "first_name": ["Prénom"],
    "email": ["Courriel 1"],
    "classes": ["Maîtrises", "Années"],
}
TEACHER_MANDATORY_FIELDS = ("first_name", "last_name", "classes")
TEACHER_COL_NAME_TO_FIELD = {}
for _field, _col_names in TEACHER_CORRESPONDANCE_DICT.items():
    for _col_name in _col_names:
        TEACHER_COL_NAME_TO_FIELD.setdefault(_col_name, _field)

YEARS_MINMAX = SchoolYear.objects.all().aggregate(Min("year"), Max("year"))
ALL_YEARS = list(range(YEARS_MINMAX["year__min"], YEARS_MINMAX["year__max"] + 1))


def load_teachers(filelike, building=None):  # noqa: CCR001
    try:
        xls_book = load_workbook(filelike)
        sheet = xls_book.active
        header_row = [cell.value for cell in sheet[1]]
        for field in TEACHER_MANDATORY_FIELDS:
            if not any(col_name in header_row for col_name in TEACHER_CORRESPONDANCE_DICT[field]):
                raise ValueError(_("Missing mandatory field: %s") % field)
    except (ValueError, KeyError) as exc:
        raise ValueError(_("File format is unreadable")) from exc

    nb_created = 0
    nb_updated = 0
    nb_skipped = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = dict(zip(header_row, row))
        translated = {}
        for key, val in values.items():
            field = TEACHER_COL_NAME_TO_FIELD.get(key)
            if field and field != "classes":
                translated[field] = val

        classes = None
        for col_name in TEACHER_CORRESPONDANCE_DICT["classes"]:
            if values.get(col_name):
                classes = values[col_name]
                break
        if not classes:
            nb_skipped += 1
            continue

        number = translated.pop("number", None)
        if number:
            teacher, created = Teacher.objects.update_or_create(number=number, defaults=translated)
        elif translated.get("first_name") and translated.get("last_name"):
            teacher, created = Teacher.objects.update_or_create(
                first_name=translated["first_name"],
                last_name=translated["last_name"],
                defaults=translated,
            )
        else:
            nb_skipped += 1
            continue
        if building:
            teacher.buildings.add(building)
        teacher.years.clear()
        if created:
            nb_created += 1
        else:
            nb_updated += 1

        years = set()
        for classes_part in classes.split(","):
            match = re.match(r"(\d+)(?:-(\d+))?[a-zA-Z]+\d?/?\w*", classes_part.strip())
            if not match:
                # ACC or DES
                years = years.union(ALL_YEARS)
            else:
                years = years.union([int(year) for year in match.groups() if year is not None])

        for year in years:
            if year is None:
                continue
            try:
                school_year = SchoolYear.objects.get(year=year)
                teacher.years.add(school_year)
            except SchoolYear.DoesNotExist:
                continue

    return nb_created, nb_updated, nb_skipped
