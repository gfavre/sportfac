import logging
import re
from datetime import datetime

from django.conf import settings
from django.utils.translation import gettext as _

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from profiles.models import School, SchoolYear
from registrations.models import Child


logger = logging.getLogger(__name__)

id_field = "id_lagapeo"
if settings.KEPCHUP_LOOKUP_AVS:
    id_field = "avs"


CHILD_MANDATORY_FIELDS = (
    id_field,
    "first_name",
    "last_name",
    "sex",
    "birth_date",
)
CORRESPONDANCE_DICT = {
    "id_lagapeo": ["ID LAGAPEO", "ELEVE_ID_PERS", "NO"],
    "first_name": ["Prénom", "ELEVE_PRENOM", "PRENOM"],
    "last_name": ["Nom", "ELEVE_NOM", "NOM"],
    "birth_date": ["Date de naissance", "ELEVE_DATE_NAISS", "DAT_NAISSANCE"],
    "sex": ["Genre", "ELEVE_GENRE", "SEXE"],
    "avs": ["AVS", "N_AVS"],
    "school_year": ["Année", "ANNEE", "ELEVE_ANNEE_SCOL", "Année2"],
    "nationality": ["Nationalité"],
    "language": ["Langue maternelle", "LANGUE MATERNELLE"],
    "school": ["Etablissement", "ETABLISSEMENT", "ETABLISSEMENT_NOM"],
    "is_blacklisted": ["Blacklist", "BLACKLIST", "Balcklist"],
    "marked_up_price": ["Prix majoré"],
}

col_name_to_field = {}
for k, v in CORRESPONDANCE_DICT.items():
    for x in v:
        col_name_to_field.setdefault(x, k)


class ChildParser:
    def __init__(self):
        self.schoolyears = {year.year: year for year in SchoolYear.objects.all()}
        self.schools = {school.code: school for school in School.objects.all()}
        self.fields_dict = {
            "id_lagapeo": self.parse_id_lagapeo,
            "first_name": lambda first_name: first_name,
            "last_name": lambda last_name: last_name,
            "birth_date": self.parse_birth_date,
            "sex": self.parse_sex,
            "nationality": self.parse_nationality,
            "language": self.parse_language,
            "school_year": self.parse_school_year,
            "school": self.parse_school,
            "is_blacklisted": self.parse_blacklist,
            "marked_up_price": self.parse_marked_up_price,
            "avs": lambda avs: avs,
        }

    @staticmethod
    def parse_id_lagapeo(value):
        try:
            return int(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def parse_blacklist(value):
        if not value or value in (0, "0", "FALSE", "Non", "NON"):
            return False
        return True

    @staticmethod
    def parse_marked_up_price(value):
        if value and value in (1, "1", "TRUE", "Oui", "OUI"):
            return True
        return False

    @staticmethod
    def parse_sex(value):
        if value in ("G", "h"):
            return Child.SEX.M
        return Child.SEX.F

    @staticmethod
    def parse_birth_date(value):
        try:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, str):
                return datetime.strptime(value, "%d.%m.%Y").date()
            return from_excel(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def parse_nationality(value):
        if value in ("Suisse", "CH"):
            return Child.NATIONALITY.CH
        if value == "Liechtenstein":
            return Child.NATIONALITY.FL
        return Child.NATIONALITY.DIV

    @staticmethod
    def parse_language(value):
        if value in ("Français", "F"):
            return Child.LANGUAGE.F
        if value == "Italien":
            return Child.LANGUAGE.I
        if value == "Allemand":
            return Child.LANGUAGE.D
        if value == "Anglais":
            return Child.LANGUAGE.E
        return Child.LANGUAGE.F

    def parse_school(self, value):
        return self.schools.get(value, None)

    def parse_school_year(self, value):
        if isinstance(value, str):
            try:
                match = re.match(r"\s*(\d+)\s?\w*.*", value)
                if not match:
                    logger.debug(f"Year not parsed: {value}")
                    return None
                value = int(match.group(1))
            except TypeError:
                return self.schoolyears.get(value, None)
            except (ValueError, IndexError):
                return None
        elif isinstance(value, float):
            value = int(value)

        year = self.schoolyears.get(value, None)
        if not year:
            logger.debug(f"no corresponding year found: {value}")
        return year

    def parse(self, row):
        out = {}
        for key, val in row.items():
            translitterated_key = col_name_to_field.get(key, None)
            try:
                translitterated_value = self.fields_dict.get(translitterated_key, lambda notfound: None)(val)
                if translitterated_value:
                    out[translitterated_key] = translitterated_value
            except Exception as exc:
                logger.warning(f"{exc}: Could not parse key={key}, value={val}")
                continue
        return out


def check_children_load_format(filelike):
    try:
        xls_book = load_workbook(filelike)
        sheet = xls_book.active
        header_row = [cell.value for cell in sheet[1]]
        for field in CHILD_MANDATORY_FIELDS:
            if not any(set(CORRESPONDANCE_DICT[field]).intersection(set(header_row))):
                raise ValueError(_("Missing mandatory field: {field}").format(field=field))

    except KeyError as exc:
        raise ValueError(_("File format is unreadable")) from exc


def load_children(filelike):
    try:
        xls_book = load_workbook(filelike)
        sheet = xls_book.active
        header_row = [cell.value for cell in sheet[1]]
        for field in CHILD_MANDATORY_FIELDS:
            if not any(set(CORRESPONDANCE_DICT[field]).intersection(set(header_row))):
                raise ValueError(f"Missing mandatory field: {field}")

    except (ValueError, KeyError) as exc:
        raise ValueError(_("File format is unreadable")) from exc
    nb_created = 0
    nb_updated = 0
    parser = ChildParser()
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        real_row_number = row_number + 2
        values = dict(zip(header_row, row))
        try:
            parsed = parser.parse(values)
        except Exception as exc:
            logger.warning(f"{exc}: Could not parse row={real_row_number}, values={values}")
            continue
        logger.debug(real_row_number, parsed)
        if not parsed.get(id_field):
            continue
        id_value = parsed.pop(id_field)
        if not parsed.get("birth_date"):
            logger.warning(f"{id_value}: Could not add, missing birth date")
            continue
        attrs = {id_field: id_value, "defaults": parsed}
        child, created = Child.objects.update_or_create(**attrs)
        if created:
            nb_created += 1
        else:
            nb_updated += 1
    return nb_created, nb_updated
